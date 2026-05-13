#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
zyoumukiroku.xlsx から列幅・行高・罫線・フォント・背景（テーマ色解決含む）を抽出し、
要約 JSON を生成する（テンプレの検査・差分確認用）。

出力:
  - src/lib/jommu-excel-spec.compact.json … 要約（Git 用・小容量）

乗務記録簿の印刷レイアウトは templates/jommu-zyoumukiroku.xlsx と src/lib/jommu-excel-fill.ts を参照。

使い方:
  python scripts/extract_jommu_excel_spec.py [入力.xlsx]
"""

from __future__ import annotations

import json
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries


def _rgb(color: Color | None) -> str | None:
    if color is None or color.type == "auto":
        return None
    if color.type == "rgb" and color.rgb:
        s = str(color.rgb)
        if len(s) == 8:
            return "#" + s[2:].upper()
        if len(s) == 6:
            return "#" + s.upper()
        return s
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}+{color.tint}"
    return str(color)


def read_theme_palette(xlsx: Path) -> list[str | None]:
    """clrScheme の sRGB（dk1/lt1 の sysClr lastClr 含む）を順に返す。"""
    with zipfile.ZipFile(xlsx) as z:
        root = ET.fromstring(z.read("xl/theme/theme1.xml"))
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", ns)
    if scheme is None:
        return []
    out: list[str | None] = []
    for child in list(scheme):
        srgb = child.find("a:srgbClr", ns)
        if srgb is not None and srgb.get("val"):
            out.append("#" + srgb.get("val", "").upper())
            continue
        sys_el = child.find("a:sysClr", ns)
        if sys_el is not None and sys_el.get("lastClr"):
            out.append("#" + sys_el.get("lastClr", "").upper())
            continue
        out.append(None)
    return out


def apply_tint_to_rgb(hex6: str, tint: float) -> str:
    """Office の theme tint（-1..1）を RGB に近似。"""
    h = hex6.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if tint < 0:
        factor = 1.0 + tint
        r = int(round(r * factor))
        g = int(round(g * factor))
        b = int(round(b * factor))
    else:
        r = int(round(r * (1.0 - tint) + 255.0 * tint))
        g = int(round(g * (1.0 - tint) + 255.0 * tint))
        b = int(round(b * (1.0 - tint) + 255.0 * tint))
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    return f"#{r:02X}{g:02X}{b:02X}"


def resolve_fill_hex(fg: Color | None, palette: list[str | None]) -> str | None:
    if fg is None or fg.type != "theme" or fg.theme is None:
        return _rgb(fg) if fg else None
    idx = int(fg.theme)
    base = palette[idx] if 0 <= idx < len(palette) else None
    if base is None or not base.startswith("#"):
        return None
    tint = float(fg.tint or 0.0)
    if abs(tint) < 1e-9:
        return base
    return apply_tint_to_rgb(base[1:], tint)


def excel_col_width_to_px(width: float | None, default: float = 1.625) -> float:
    w = default if width is None else float(width)
    return round((w + 0.71) * 7, 3)


def _side(b: Any, name: str) -> dict[str, Any] | None:
    if b is None:
        return None
    side = getattr(b, name, None)
    if side is None or side.style is None:
        return None
    return {"style": side.style, "color": _rgb(side.color) if side.color else None}


def _border(b: Any) -> dict[str, Any]:
    if b is None:
        return {}
    return {
        "left": _side(b, "left"),
        "right": _side(b, "right"),
        "top": _side(b, "top"),
        "bottom": _side(b, "bottom"),
    }


def _font(f: Any) -> dict[str, Any]:
    if f is None:
        return {}
    return {
        "name": f.name,
        "size": f.sz,
        "bold": bool(f.b),
        "color": _rgb(f.color) if f.color else None,
    }


def _fill(fl: Any, palette: list[str | None]) -> dict[str, Any]:
    if fl is None or fl.fill_type is None:
        return {"type": None, "resolvedHex": None}
    out: dict[str, Any] = {"type": fl.fill_type}
    if fl.fill_type == "solid" and fl.fgColor:
        raw = _rgb(fl.fgColor)
        out["raw"] = raw
        out["resolvedHex"] = resolve_fill_hex(fl.fgColor, palette) if fl.fgColor.type == "theme" else raw
    return out


def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def meaningful_cell(c: Cell) -> bool:
    if c.value is not None and str(c.value).strip() != "":
        return True
    if c.fill and c.fill.fill_type == "solid":
        return True
    b = c.border
    if b and any(
        [
            b.left and b.left.style,
            b.right and b.right.style,
            b.top and b.top.style,
            b.bottom and b.bottom.style,
        ]
    ):
        return True
    return False


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    default_in = root.parent / "新しいフォルダー" / "zyoumukiroku.xlsx"
    alt_in = root / "新しいフォルダー" / "zyoumukiroku.xlsx"
    in_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_in
    if not in_path.is_file() and alt_in.is_file():
        in_path = alt_in
    if not in_path.is_file():
        print(f"ERROR: not found: {in_path}", file=sys.stderr)
        return 1

    palette = read_theme_palette(in_path)
    wb = load_workbook(in_path, data_only=False)
    ws = wb[wb.sheetnames[0]]

    merged: list[dict[str, Any]] = []
    for r in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(r))
        merged.append({"range": str(r), "minRow": min_row, "maxRow": max_row, "minCol": min_col, "maxCol": max_col})

    col_widths: dict[str, dict[str, Any]] = {}
    for letter, dim in ws.column_dimensions.items():
        if letter and dim.width is not None:
            w = dim.width
            col_widths[letter] = {"excelWidth": w, "approxPx": excel_col_width_to_px(w)}

    row_heights: dict[str, float] = {}
    for idx, dim in ws.row_dimensions.items():
        if dim.height is not None:
            row_heights[str(idx)] = round(float(dim.height), 3)

    cells: dict[str, Any] = {}
    for row in range(1, (ws.max_row or 0) + 1):
        for col in range(1, (ws.max_column or 0) + 1):
            c = ws.cell(row=row, column=col)
            if not meaningful_cell(c):
                continue
            addr = cell_addr(row, col)
            v = c.value
            if v is not None and not isinstance(v, (str, int, float, bool)):
                v = str(v)
            fl = _fill(c.fill, palette)
            cells[addr] = {
                "row": row,
                "col": col,
                "value": v,
                "font": _font(c.font),
                "fill": fl,
                "border": _border(c.border),
            }

    def px_for_col(ci: int) -> float:
        letter = get_column_letter(ci)
        dim = col_widths.get(letter)
        return excel_col_width_to_px(dim["excelWidth"] if dim else None)

    spans = [
        ("A", "A"),
        ("B", "J"),
        ("K", "T"),
        ("U", "AC"),
        ("AD", "AH"),
        ("AI", "AP"),
        ("AQ", "AZ"),
        ("BA", "BE"),
        ("BF", "BK"),
        ("BL", "BQ"),
        ("BR", "BW"),
        ("BX", "CB"),
    ]
    col_px: list[float] = []
    for a, b in spans:
        s = sum(px_for_col(c) for c in range(column_index_from_string(a), column_index_from_string(b) + 1))
        col_px.append(round(s, 3))
    tot = sum(col_px) or 1.0
    col_frac = [round(x / tot, 6) for x in col_px]

    a6 = ws["A6"]
    header_fill = _fill(a6.fill, palette).get("resolvedHex") or "#FCE4D6"
    a2 = ws["A2"]
    title_font = _font(a2.font)
    bg1 = ws["BG1"]
    retention_font = _font(bg1.font)

    compact: dict[str, Any] = {
        "sourceFile": str(in_path.resolve()),
        "sheet": ws.title,
        "palette": palette,
        "headerFillResolved": header_fill,
        "titleFont": title_font,
        "retentionFont": retention_font,
        "mainTableColWidthsPx": col_px,
        "mainTableColFractions": col_frac,
        "rowHeightsPt": row_heights,
        "mergedRangeCount": len(merged),
        "cellsSampleCount": len(cells),
    }
    compact_path = root / "src" / "lib" / "jommu-excel-spec.compact.json"
    compact_path.write_text(json.dumps(compact, ensure_ascii=False, indent=2), encoding="utf-8")

    # 旧フル JSON は巨大なので削除（存在すれば）
    legacy = root / "src" / "lib" / "jommu-excel-spec.json"
    if legacy.is_file():
        legacy.unlink()

    print(f"Wrote {compact_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
